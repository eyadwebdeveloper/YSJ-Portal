from flask import Flask, request, render_template
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

def write_to_excel(data):
    # Load the existing workbook or create a new one if it doesn't exist
    try:
        wb = load_workbook('form_data.xlsx')
    except FileNotFoundError:
        wb = Workbook()
        wb.save('form_data.xlsx')
    ws = wb.active

    # Append form data to the Excel file
    ws.append([data['name'], data['phone'], data['birth'], data['gender'], data['school'], data['grade'], data['nationality'], data['email']])

    # Save the workbook
    wb.save('form_data.xlsx')

@app.route('/')
def index():
    return render_template('info.html')

@app.route('/submit_form', methods=['POST'])
def submit_form():
    if request.method == 'POST':
        form_data = {
            'name': request.form['name'],
            'phone': request.form['phone'],
            'birth': request.form['birth'],
            'gender': request.form['gender'],
            'school': request.form['school'],
            'grade': request.form['grade'],
            'nationality': request.form['nationality'],
            'email': request.form['email']
        }

        # Write form data to Excel
        write_to_excel(form_data)

        return 'Form submitted successfully!'

if __name__ == '__main__':
    app.run(debug=True)
